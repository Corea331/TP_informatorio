from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

def crear_archivo_excel(empleado, empleador, datos_pago, nombre_mes, año):
    nombre_archivo = f"Pago_{empleado.nombre}_{empleado.cuil}.xlsx"

    # si el arhivo ya existe, lo abre; si no, lo crea
    if os.path.exists(nombre_archivo):
        libro = load_workbook(nombre_archivo)
    else:
        libro = Workbook()

    # crea una hoja para guardar los datos personales del empleado y el empleador
    ws_datos = libro.active
    ws_datos.title = "Datos Personales"

    # Agrega los datos del empleado
    ws_datos.append(["Nombre", empleado.nombre])
    ws_datos.append(["Apellido", empleado.apellido])
    ws_datos.append(["Cuil", empleado.cuil])
    ws_datos.append(["Fecha de Ingreso", empleado.fecha_ingreso])
    ws_datos.append(["antiguedad", empleado.antiguedad])
    ws_datos.append(["Domicilio", empleado.domicilio])
    ws_datos.append(["Teléfono", empleado.telefono])
    ws_datos.append(["Email", empleado.email])
    ws_datos.append(["Modalidad", empleado.modalidad])


    # Agrega los datos del empleador
    ws_datos.append(["Nombre del Empleador", empleador.nombre])
    ws_datos.append(["Apellido del Empleador", empleador.apellido])
    ws_datos.append(["Cuil del Empleador", empleador.cuil])
    ws_datos.append(["Domicilio del Empleador", empleador.domicilio])
    ws_datos.append(["Teléfono del Empleador", empleador.telefono])
    ws_datos.append(["Email del Empleador", empleador.email])

    # crea una hoja para guardar los datos de pago del mes
    nombre_hoja = f"{nombre_mes}_{año}"
    if nombre_hoja in libro.sheetnames:
        del libro[nombre_hoja]
    ws_pago = libro.create_sheet(title=nombre_hoja)

    # guardar los datos del mes a pagar
    ws_pago.append(["concepto", "valor"])
    for clave, valor in datos_pago.items():
        ws_pago.append([clave.replace("_", " ").capitalize(), valor])

    # Ajustar el ancho de las columnas
    for col in ws_pago.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_pago.column_dimensions[col_letter].width = max_length + 2

    # Guardar el archivo
    libro.save(nombre_archivo)
    print(f"Archivo {nombre_archivo} guardado exitosamente.")